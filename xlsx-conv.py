from openpyxl import load_workbook
from time import strftime
from itertools import islice
from io import StringIO
import argparse
import csv
import os
import sys


# ---
# GET ARGUMENTS
# ---
# see https://docs.python.org/3/howto/argparse.html 

parser = argparse.ArgumentParser(description='Convert XLSX file to DSV using openpyxl')
parser.add_argument('-i','--input', help='Path to XLSX/XLSM/XLTX/XLTM file or path to tab separated, UTF-8 encoded TXT file with a column named "input" and optional columns "outputDir" and "prefix". The TXT file should have values for the supplied columns on each line.', required=True)
parser.add_argument('-o','--outputDir', help='Path to output directory')
parser.add_argument('--col_index', help='Generate column indices (c1,c2,etc) as first line in output', action="store_true")
parser.add_argument('--delimiter', help='Delimiter used in output, defaults to ,', choices=[',', ';', '|', 'tab'], default=',')
parser.add_argument('--encoding', help='Output encoding, defaults to utf-8. Warning: ascii and latin-1 codecs cannot encode all characters that may be in Excel file and conversion will fail if the Excel file contains those characters.', choices=['ascii', 'latin-1', 'utf-8', 'utf-16'], default='utf-8')
parser.add_argument('--extension', help='Extension of output, defaults to csv', default='csv')
parser.add_argument('--linebreak_replacement', help='Replace linebreaks in cells by replacement string')
parser.add_argument('--max_cols', type=int, help="Maximum number of columns", default=-1)
parser.add_argument('--noprefix', help='Do not prefix ouput with workbook name', action="store_true")
parser.add_argument('--prefix', help='Use specified prefix instead of prefixing output with workbook name')
parser.add_argument('--row_index', help='Write row numbers as first column in output', action="store_true")
parser.add_argument('--quotechar', help='One-character string used to quote fields containing special characters', default='"')
parser.add_argument('--quoting', help='Controls field quoting, defaults to MINIMAL', choices=['ALL', 'MINIMAL', 'NONE', 'NONNUMERIC'], default='MINIMAL')
parser.add_argument('--sheet', help='Name of sheet to convert')
parser.add_argument('--sheetnames', help='Do not convert input, but echo sheetnames', action="store_true")
parser.add_argument('--version', action='version', version="%(prog)s 1.4.0-beta")
args = parser.parse_args()

# ---
# CONSTANTS
# ---
# used as keys in arg dictionary

# arguments taken either from commandline or TXT input
INPUT_PATH = "inputpath"
OUTPUT_DIR = "outputdir"
PREFIX = "prefix"
SHEET = "sheet"

# shared arguments
# these will be the same for each row in inputList
COL_INDEX = 'col_index'
DELIMITER = 'delimiter'
ENCODING = "encoding"
EXTENSION = "extension"
LINEBREAK_REPLACEMENT = 'linebreak_replacement'
MAX_COLS = 'max_cols'
NO_PREFIX = "noprefix"
ROW_INDEX = "row_index"
QUOTECHAR = "quotechar"
QUOTING = "quoting"
SHEETNAMES = "sheetnames"

# derived
INPUT_FILE = "inputfile"
INPUT_BASE_FN = "inputbasefn"

# ---
# ARGUMENT HANDLING
# ---

# In the main program we always want to iterate over a list of dictionaries with the conversion arguments
#
#  - If the input is an OOXML Excel file (xlsx/xlsm/xltx/xltm) then the list will have only one row
#  - If the input a a TXT file then we fill the list with entries found in that txt file
#  - If no outputDir is specified then outputDir wil be the same as the inputDir
#
# Input has to be a file otherwise exit immediately

inputPath = args.input
inputBaseFn, inputExt = os.path.splitext(inputPath)

validInputExt = { ".xlsx", ".xlsm", ".xltx", ".xltm",".txt"}

if inputExt.lower() not in validInputExt: 
    print('Error: Expecting xlsx/xlsm/xltx/xltm or txt input\n')
    sys.exit(1)

inputList = []

if inputExt.lower() == ".txt":
    if not os.path.isfile(inputPath):
        print('Error: No such file', inputPath, '\n')
        sys.exit(1)
    with open(inputPath) as tsv_file:
        csv_reader = csv.reader(tsv_file, delimiter='\t')
        headers = next(csv_reader,[])
        lowerCaseHeaders = [h.lower() for h in headers]
        indexOutputDir = -1 # outputDir is optional
        indexPrefix = -1 # prefix is optional
        indexSheet = -1 # sheet is optional
        if "input" in lowerCaseHeaders:
            indexInput = lowerCaseHeaders.index("input") # index method returns value error if value is not in list!
        else:
            print('Error: Unable to find header named input in txt file\n')
            sys.exit(1)
        if OUTPUT_DIR in lowerCaseHeaders:
            indexOutputDir = lowerCaseHeaders.index(OUTPUT_DIR)
        if PREFIX in lowerCaseHeaders:
            indexPrefix = lowerCaseHeaders.index(PREFIX)
        if SHEET in lowerCaseHeaders:
            indexSheet = lowerCaseHeaders.index(SHEET)
        for row in csv_reader:
            inputDict = {}
            inputDict[INPUT_PATH] = row[indexInput]
            # outputDir
            if indexOutputDir >= 0:
                inputDict[OUTPUT_DIR] = row[indexOutputDir]
            else:
                inputDict[OUTPUT_DIR] = None
            # prefix
            if indexPrefix >= 0:
                inputDict[PREFIX] = row[indexPrefix]
            else:
                inputDict[PREFIX] = args.prefix
            # sheet
            if indexSheet >= 0:
                inputDict[SHEET] = row[indexSheet]
            else:
                inputDict[SHEET] = None
            inputList.append(inputDict)

if inputExt.lower() != ".txt":
    inputDict = {}
    inputDict[INPUT_PATH] = inputPath
    inputDict[OUTPUT_DIR] = args.outputDir
    inputDict[PREFIX] = args.prefix
    inputDict[SHEET] = args.sheet
    inputList.append(inputDict)

# ---
# Define functions we use in MAIN 
# These need to be defined up front
# ---

#
# convert sheet function
# 
def convertSheet(ws,outputPath,argDict):
    # quote style
    quoting = argDict[QUOTING]
    quoteStyle = csv.QUOTE_MINIMAL
    if quoting == "ALL":
      quoteStyle = csv.QUOTE_ALL
    elif quoting == "NONE":
      quoteStyle = csv.QUOTE_NONE
    elif quoting == "NONNUMERIC":
      quoteStyle = csv.QUOTE_NONNUMERIC

    outputEncoding = argDict[ENCODING]
    outputDelimiter = argDict[DELIMITER]
    outputQuoteChar = argDict[QUOTECHAR]
    linebreakReplacement = argDict[LINEBREAK_REPLACEMENT]
    colIndex = argDict[COL_INDEX]
    rowIndex = argDict[ROW_INDEX]
    maxColumns = argDict[MAX_COLS]
    
    # First check if there are rows in ws.rows 
    # The sheet may be empty
    first_row_slice = list(islice(ws.rows,1))
    if len(first_row_slice) == 0:
        print(strftime("%Y-%m-%d %H:%M:%S"), "- Skipping conversion of sheet", ws.title, "since it is empty.")
        return

    with open(outputPath, 'w', encoding=outputEncoding) as f:
        c = csv.writer(f, lineterminator='\n', delimiter=outputDelimiter, quotechar=outputQuoteChar, quoting=quoteStyle)
        
        first_row = first_row_slice[0]
        numcols = len(first_row)
        if maxColumns > -1:
            if numcols > maxColumns:
                numcols = maxColumns
                print(strftime("%Y-%m-%d %H:%M:%S"), "- Limiting output to", numcols, "columns")

        # We iterate over range(numcols), e.g. range(3) will give [0,1,2]
        if colIndex == True:
            first_row =  first_row_slice[0]
            col_index = []
            if rowIndex == True:
                col_index.append("c0") # if row_index is also set then include additional column
            for i in range(numcols):
                c_val = "c" + str(i+1)
                col_index.append(c_val)
            c.writerow(col_index)

        print(strftime("%Y-%m-%d %H:%M:%S"), "- Outputting converted sheet to", outputPath)
        for index, row in enumerate(ws.rows):
            values = []
            if rowIndex == True:
                values.append(index+1)
            for colnum in range(numcols):
                cell = row[colnum]
                value = cell.value
                if linebreakReplacement is not None and isinstance(value, str):
                    value = value.replace('\r\n', linebreakReplacement).replace('\n', linebreakReplacement).replace('\r', linebreakReplacement)
                values.append(value)
            c.writerow(values)

#
# convertWorkbook function
#
def convertWorkbook(argDict):
    inputPath = argDict[INPUT_PATH]
    print(strftime("%Y-%m-%d %H:%M:%S"), "- Converting", inputPath)
    try:
        wb = load_workbook(filename=inputPath, read_only=True, data_only=True)
    except Exception as e:
        print('Error: Failed to load workbook from', inputPath, '\n')
        print(e)
        sys.exit(1)

    ws_names = wb.sheetnames
    if argDict[SHEET]:
        if(argDict[SHEET] in ws_names):
            print(strftime("%Y-%m-%d %H:%M:%S"), "- Only extracting sheet", argDict[SHEET])
            ws_names = [argDict[SHEET]]
        else:
            print('Error: Cannot find sheet ', argDict[SHEET])
            sys.exit(1)
    outputPrefix = argDict[INPUT_BASE_FN] + '.'
    if argDict[PREFIX]:
        outputPrefix = argDict[PREFIX] + '.' # override with custom prefix
    if argDict[NO_PREFIX]:
        outputPrefix = '' # noprefix takes precedence over customPrefix
    for ws_name in ws_names:
        ws = wb[ws_name] # ws is an IterableWorksheet
        outputPath = outputDir + os.sep + outputPrefix + ws_name + '.' + argDict[EXTENSION]
        try:
            convertSheet(ws,outputPath,argDict)
        except Exception as e:
            print('Error: Failed to convert sheet\n')
            print(e)
            sys.exit(1) 

#
# listSheetnames function
#
def listSheetnames(argDict):
    inputPath = argDict[INPUT_PATH]
    try:
        wb = load_workbook(filename=inputPath, read_only=True)
    except Exception as e:
        print('Error: Failed to load workbook from', inputPath, '\n')
        print(e)
        sys.exit(1)

    ws_names = wb.sheetnames
    # forward slash is not allowed in Excel sheetname nor in Linux or Windows filename
    for ws_name in ws_names:
        l = [ws_name, argDict[INPUT_PATH]]
        line = StringIO()
        writer = csv.writer(line,lineterminator='\n', delimiter=",")
        writer.writerow(l)
        csvcontent = line.getvalue()
        print(csvcontent, end = '')

# ---
# MAIN
# --

if args.sheetnames:
    print("Sheet,Input")

for d in inputList:
    # input
    realInputPath = os.path.realpath(d[INPUT_PATH])
    if not os.path.isfile(realInputPath):
        print('Error: No such file ', realInputPath, '\n')
        sys.exit(1)
    inputDir, inputFile = os.path.split(realInputPath)
    inputBaseFn, inputExt = os.path.splitext(inputFile)

    d[INPUT_FILE] = inputFile
    d[INPUT_BASE_FN] = inputBaseFn

    if args.sheetnames:
        listSheetnames(d)
        continue

    # output
    outputDir = d[OUTPUT_DIR]
    if not outputDir:
       outputDir = inputDir
    elif not os.path.isdir(outputDir):
       print('Error: No such directory ', outputDir, '\n')
       sys.exit(1)

    d[OUTPUT_DIR] = outputDir

    # shared args
    d[COL_INDEX] = args.col_index
    d[DELIMITER] = args.delimiter
    d[ENCODING] = args.encoding
    d[EXTENSION] = args.extension
    d[LINEBREAK_REPLACEMENT] = args.linebreak_replacement
    d[MAX_COLS] = args.max_cols
    d[NO_PREFIX] = args.noprefix
    d[ROW_INDEX] = args.row_index
    d[QUOTECHAR] = args.quotechar
    d[QUOTING] = args.quoting

    # call convert workbook
    convertWorkbook(d)

if not(args.sheetnames):
    print(strftime("%Y-%m-%d %H:%M:%S"), "- Finished!")